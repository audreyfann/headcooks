"""
Parses a multi-sheet xlsx workbook into a flat list of ingredient rows.

Expected format per sheet (one sheet per cook/event):
    Name: <cook name>              (optional, row 1-6 area)
    Title: <event title>           (optional)
    Date: <event date>             (optional)

    <Dish Name>
    Recipe Instructions: ...
    Ingredient | Quantity | Notes | Status (...) | Supplier | Arrived (...)
    <ingredient row>
    <ingredient row>
    ...
    <Next Dish Name>
    ...

The header row can appear multiple times per sheet (once per dish/recipe
block). Column order and exact header wording can vary sheet-to-sheet --
columns are matched by keyword (see config.HEADER_KEYWORDS), not by fixed
position.
"""

from __future__ import annotations

import openpyxl
import pandas as pd

from . import config


def _cell(row, i):
    return row[i] if i < len(row) else None


def _detect_header(row: tuple) -> dict | None:
    """If this row looks like a table header (contains 'Quantity'), return a
    mapping of column index -> field name. Otherwise return None."""
    if not any(isinstance(c, str) and c.strip() == "Quantity" for c in row):
        return None

    header_map = {}
    for idx, c in enumerate(row):
        if not isinstance(c, str) or not c.strip():
            continue
        key = c.strip().lower()
        if key == "`":  # some sheets use a stray backtick as the ingredient column header
            header_map[idx] = "ingredient"
            continue
        for keyword, field in config.HEADER_KEYWORDS.items():
            if keyword in key:
                header_map[idx] = field
                break
    return header_map


def _extract_cook_name(ws, sheet_name: str) -> str:
    """Look for a 'Name:' row near the top of the sheet; fall back to sheet name."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row and row[0] and isinstance(row[0], str) and row[0].strip().lower().startswith("name"):
            if len(row) > 1 and row[1]:
                return row[1]
    return sheet_name


def extract_sheet(ws, sheet_name: str) -> list[dict]:
    """Extract all ingredient rows from a single worksheet."""
    cook_name = _extract_cook_name(ws, sheet_name)

    current_dish = None
    header_map: dict | None = None
    records: list[dict] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        detected = _detect_header(row)
        if detected:
            header_map = detected
            continue

        c1, c2 = _cell(row, 1), _cell(row, 2)

        if header_map is None:
            # Not currently inside an ingredient table -- this might be a dish title row
            if isinstance(c1, str) and c1.strip() and c2 is None and not c1.strip().lower().startswith("recipe instructions"):
                current_dish = c1.strip()
            continue

        ing_idx = [k for k, v in header_map.items() if v == "ingredient"]
        qty_idx = [k for k, v in header_map.items() if v == "quantity"]
        ingredient_val = _cell(row, ing_idx[0]) if ing_idx else None
        quantity_val = _cell(row, qty_idx[0]) if qty_idx else None

        # A new dish section is starting (title row with no quantity/ingredient values)
        if isinstance(c1, str) and c1.strip() and quantity_val is None and ingredient_val is None:
            if not c1.strip().lower().startswith("recipe instructions"):
                current_dish = c1.strip()
            header_map = None
            continue

        if ingredient_val is None and quantity_val is None:
            continue  # blank row

        if isinstance(ingredient_val, str) and ingredient_val.strip().lower().startswith("recipe instructions"):
            continue

        record = {"cook": cook_name, "sheet": sheet_name, "dish": current_dish}
        for idx, field in header_map.items():
            record[field] = _cell(row, idx)
        records.append(record)

    return records


def extract_workbook(path: str, skip_sheets: set[str] | None = None) -> pd.DataFrame:
    """Extract ingredient rows from every sheet in the workbook (except skipped ones).

    Returns a DataFrame with columns: cook, sheet, dish, ingredient, quantity,
    notes, status, supplier, arrived (columns present depend on what each
    sheet's header row actually contained).
    """
    skip_sheets = skip_sheets or config.DEFAULT_SKIP_SHEETS
    wb = openpyxl.load_workbook(path, data_only=True)

    all_records = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        all_records.extend(extract_sheet(ws, sheet_name))

    df = pd.DataFrame(all_records)
    if df.empty:
        return df

    df = df[df["ingredient"].notna() & (df["ingredient"].astype(str).str.strip() != "")]

    preferred_order = ["cook", "dish", "ingredient", "quantity", "notes", "status", "supplier", "arrived", "sheet"]
    cols = [c for c in preferred_order if c in df.columns]
    return df[cols].reset_index(drop=True)
